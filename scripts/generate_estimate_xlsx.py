#!/usr/bin/env python3
"""Generate a Chinese software estimation workbook from validated JSON."""

from __future__ import annotations

import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from validate_estimate_data import load_json, validate_estimate


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
MONEY_HEADERS = {"内部估算金额", "客户报价金额", "报价金额"}
DAY_HEADERS = {
    "基础总人天",
    "Buff总人天",
    "Buff后总人天",
    "产品人天",
    "UI人天",
    "iOS人天",
    "Android人天",
    "管理后台Web人天",
    "H5人天",
    "小程序人天",
    "PC Web人天",
    "前端小计人天",
    "后端人天",
    "测试人天",
    "项目管理人天",
    "基础合计人天",
    "Buff人天",
    "最终合计人天",
    "基准人天",
    "最终基础人天",
}


SUMMARY_FIELDS = [
    ("项目名称", ("project", "name")),
    ("业务领域", ("project", "business_domain")),
    ("估算范围", ("project", "scope")),
    ("涉及端", ("project", "included_terminals")),
    ("估算模式", ("project", "estimation_mode")),
    ("基础总人天", ("summary", "base_total_days")),
    ("Buff总人天", ("summary", "buff_total_days")),
    ("Buff后总人天", ("summary", "final_total_days")),
    ("平均Buff比例", ("summary", "average_buff_percent")),
    ("Buff原因汇总", ("summary", "buff_reason_summary")),
    ("内部估算金额", ("summary", "internal_estimated_amount")),
    ("客户报价金额", ("summary", "customer_quotation_amount")),
    ("报价系数", ("pricing", "quotation_coefficient")),
    ("不含范围", ("project", "excluded_scope")),
    ("关键假设", ("project", "key_assumptions")),
]

FUNCTION_HEADERS = [
    ("一级模块", "module"),
    ("二级功能", "function"),
    ("功能说明", "description"),
    ("涉及端", "included_terminals"),
    ("覆盖方式", "coverage_mode"),
    ("覆盖依据", "coverage_evidence"),
    ("复杂度", "complexity"),
    ("风险等级", "risk_level"),
    ("产品人天", "product_days"),
    ("UI人天", "ui_days"),
    ("iOS人天", "ios_days"),
    ("Android人天", "android_days"),
    ("管理后台Web人天", "web_admin_days"),
    ("H5人天", "h5_days"),
    ("小程序人天", "mini_program_days"),
    ("PC Web人天", "pc_web_days"),
    ("前端小计人天", "frontend_subtotal_days"),
    ("后端人天", "backend_days"),
    ("测试人天", "test_days"),
    ("项目管理人天", "project_management_days"),
    ("基础合计人天", "base_total_days"),
    ("Buff比例", "buff_percent"),
    ("Buff人天", "buff_days"),
    ("Buff原因", "buff_reason"),
    ("最终合计人天", "final_total_days"),
    ("报价金额", "quotation_amount"),
    ("备注", "notes"),
]

DETAIL_HEADERS = [
    ("一级模块", "module"),
    ("二级功能", "function"),
    ("实施项类型", "item_type"),
    ("实施项名称", "item_name"),
    ("端类型", "terminal_type"),
    ("角色", "role"),
    ("基准人天", "base_days"),
    ("覆盖系数", "coverage_coefficient"),
    ("复杂度系数", "complexity_coefficient"),
    ("风险系数", "risk_coefficient"),
    ("最终基础人天", "final_base_days"),
    ("估算依据", "estimation_evidence"),
    ("待确认点", "pending_confirmation"),
]

RISK_HEADERS = [
    ("风险类别", "category"),
    ("风险说明", "description"),
    ("影响模块", "affected_module"),
    ("风险等级", "risk_level"),
    ("对工时影响", "day_impact"),
    ("建议处理方式", "suggested_handling"),
    ("是否已计入报价", "included_in_quotation"),
]

QUESTION_HEADERS = [
    ("问题分类", "category"),
    ("关联模块", "related_module"),
    ("确认问题", "question"),
    ("为什么要问", "why_it_matters"),
    ("影响工时范围", "effort_impact_range"),
    ("默认假设", "default_assumption"),
    ("客户回答", "customer_answer"),
    ("是否影响报价", "affects_quotation"),
]


def _get_nested(data: dict[str, Any], path: tuple[str, str]) -> Any:
    parent, key = path
    return data.get(parent, {}).get(key, "")


def _format_value(value: Any) -> Any:
    if isinstance(value, list):
        return "、".join(str(item) for item in value)
    return value


def _style_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            header = ws.cell(row=1, column=cell.column).value
            if header in MONEY_HEADERS and isinstance(cell.value, (int, float)):
                cell.number_format = '¥#,##0.00'
            elif header in DAY_HEADERS and isinstance(cell.value, (int, float)):
                cell.number_format = '0.00'
            elif header and "比例" in str(header) and isinstance(cell.value, (int, float)):
                cell.number_format = '0.00%'

    for column_cells in ws.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, min(len(value), 40))
        ws.column_dimensions[column_letter].width = max(12, max_length + 2)


def _write_table(ws, headers: list[tuple[str, str]], rows: list[dict[str, Any]]) -> None:
    ws.append([header for header, _ in headers])
    for row in rows:
        ws.append([_format_value(row.get(key, "")) for _, key in headers])
    _style_sheet(ws)


def _write_summary(ws, data: dict[str, Any]) -> None:
    ws.append(["字段", "值"])
    for label, path in SUMMARY_FIELDS:
        ws.append([label, _format_value(_get_nested(data, path))])
    _style_sheet(ws)
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 80


def build_workbook(data: dict[str, Any]) -> Workbook:
    wb = Workbook()

    summary_ws = wb.active
    summary_ws.title = "报价汇总"
    _write_summary(summary_ws, data)

    functions_ws = wb.create_sheet("功能报价清单")
    _write_table(functions_ws, FUNCTION_HEADERS, data.get("functions", []))

    details_ws = wb.create_sheet("实施明细")
    _write_table(details_ws, DETAIL_HEADERS, data.get("implementation_details", []))

    risks_ws = wb.create_sheet("风险与假设")
    _write_table(risks_ws, RISK_HEADERS, data.get("risks", []))

    questions_ws = wb.create_sheet("客户确认问题")
    _write_table(questions_ws, QUESTION_HEADERS, data.get("confirmation_questions", []))

    return wb


def generate_workbook(input_path: Path, output_path: Path) -> None:
    data = load_json(input_path)
    errors = validate_estimate(data)
    if errors:
        joined = "\n".join(f"- {error}" for error in errors)
        raise ValueError(f"Invalid estimate data:\n{joined}")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = build_workbook(data)
    wb.save(output_path)


def main(argv: list[str]) -> int:
    if len(argv) != 3:
        print("Usage: generate_estimate_xlsx.py <estimate.json> <output.xlsx>", file=sys.stderr)
        return 1

    try:
        generate_workbook(Path(argv[1]), Path(argv[2]))
    except Exception as exc:
        print(f"Workbook generation failed: {exc}", file=sys.stderr)
        return 1

    print(f"Workbook generated: {argv[2]}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
